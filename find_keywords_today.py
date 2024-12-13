from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook
from datetime import datetime

# Path to the Excel file
excel_path = "date_keywords.xlsx"


def fetch_all_suggestions(driver, keyword):
    """Fetch all autocomplete suggestions for a given keyword from Google."""
    # today_date = datetime.today().strftime('%Y-%m-%d')
    # print(f"Today's date : {today_date}")

    try:
        driver.get("https://www.google.com")
        time.sleep(2)

        # Input keyword into the search box
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(keyword)
        time.sleep(2)

        # Get today's date using JavaScript
        today_date = driver.execute_script("return new Date().toISOString().slice(0, 10);")
        print(f"Today's date: {today_date}")

        # Locate all suggestions
        suggestion_elems = driver.find_elements(By.CSS_SELECTOR, "ul[role='listbox'] li span")
        suggestions = [elem.text.strip() for elem in suggestion_elems if elem.text.strip()]

        if not suggestions:
            return ["No suggestions"], [keyword]

        # Longest and shortest suggestion for each keyword
        keyword_len = len(keyword.split())
        shortest = [s for s in suggestions if len(s.split()) <= keyword_len + 1]
        longest = [s for s in suggestions if len(s.split()) > keyword_len + 1]

        if not shortest:
            shortest = [min(suggestions, key=len)]
        if not longest:
            longest = [max(suggestions, key=len)]
        return longest, shortest

    except Exception as e:
        print(f"Error while fetching suggestions for '{keyword}': {e}")
        return ["No suggestions"], [keyword]


def update_sheet(sheet, results):
    """Update the sheet with all longest and shortest suggestions for each keyword."""
    for keyword, row, longest, shortest in results:
        longest_str = ", ".join(longest) if longest else "No suggestions"
        shortest_str = ", ".join(shortest) if shortest else "No suggestions"

        # Write longest and shortest suggestions into the Excel sheet
        sheet.cell(row=row, column=4, value=longest_str)
        sheet.cell(row=row, column=5, value=shortest_str)


def validate_keywords(keywords, sheet_name):
    """Ensure keywords are valid."""
    if not keywords:
        print(f"No keywords found in sheet '{sheet_name}'.")
        return False
    print(f"Keywords extracted from '{sheet_name}': {[kw[0] for kw in keywords]}")
    return True


def process_tab(driver, book, sheet_name):
    """Process a single sheet in the Excel file."""
    try:
        sheet = book[sheet_name]
        keywords = [
            (row[0].value.strip(), row[0].row)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)
            if row[0].value
        ]
        if not validate_keywords(keywords, sheet_name):
            return

        results = []
        for keyword, row in keywords:
            print(f"Fetching suggestions for: {keyword}")
            longest, shortest = fetch_all_suggestions(driver, keyword)
            results.append((keyword, row, longest, shortest))

        update_sheet(sheet, results)
        book.save(excel_path)
        print(f"Sheet '{sheet_name}' updated successfully.")
    except Exception as e:
        print(f"Error processing tab '{sheet_name}': {e}")


def main():
    """Visit all tabs in the Excel file and process them."""
    driver = webdriver.Chrome()
    try:
        book = load_workbook(excel_path)
        for sheet_name in book.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            process_tab(driver, book, sheet_name)

        print("All tabs processed successfully.")
    finally:
        driver.quit()
        print("Browser closed.")


if __name__ == "__main__":
    main()
